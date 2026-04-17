#!/usr/bin/env python
# -*- coding: utf-8 -*-
# time: 2024/4/17
# file: tasks.py
# author: 报表定时生成任务
# QQ: 939589097
import os
import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.db.models import Count
from django.db.models.functions import TruncDate
from system.models import Users, Role, Dept, OperationLog, LoginLog
from fuadmin.celery import app
from django.conf import settings


def create_styles(workbook):
    styles = {
        'header': Font(name='微软雅黑', size=12, bold=True, color='FFFFFF'),
        'header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'header_border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        ),
        'cell': Font(name='微软雅黑', size=10),
        'cell_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'cell_border': Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        ),
    }
    return styles


def apply_header_style(cell, styles):
    cell.font = styles['header']
    cell.fill = styles['header_fill']
    cell.alignment = styles['header_alignment']
    cell.border = styles['header_border']


def apply_cell_style(cell, styles):
    cell.font = styles['cell']
    cell.alignment = styles['cell_alignment']
    cell.border = styles['cell_border']


@app.task(name="system.tasks.generate_user_report")
def generate_user_report(days: int = 30):
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    workbook = Workbook()
    styles = create_styles(workbook)
    
    ws_trend = workbook.active
    ws_trend.title = "用户增长趋势"
    
    users = (
        Users.objects
        .filter(create_datetime__gte=start_date)
        .annotate(date=TruncDate('create_datetime'))
        .values('date')
        .annotate(count=Count('id'))
        .order_by('date')
    )
    
    date_dict = {item['date']: item['count'] for item in users}
    data = []
    for i in range(days):
        current_date = (start_date + timedelta(days=i)).date()
        data.append({
            "日期": current_date.strftime("%Y-%m-%d"),
            "新增用户数": date_dict.get(current_date, 0)
        })
    
    headers = ["日期", "新增用户数"]
    for col, header in enumerate(headers, 1):
        cell = ws_trend.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_trend.column_dimensions[chr(64 + col)].width = 20
    
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_trend.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    ws_distribution = workbook.create_sheet("用户分布")
    
    distribution_data = []
    total_users = Users.objects.count()
    active_users = Users.objects.filter(status=True).count()
    inactive_users = Users.objects.filter(status=False).count()
    
    distribution_data.append({"指标": "总用户数", "数值": total_users})
    distribution_data.append({"指标": "活跃用户数", "数值": active_users})
    distribution_data.append({"指标": "非活跃用户数", "数值": inactive_users})
    
    for value, name in Users.GENDER_CHOICES:
        count = Users.objects.filter(gender=value).count()
        distribution_data.append({"指标": f"{name}用户数", "数值": count})
    
    headers = ["指标", "数值"]
    for col, header in enumerate(headers, 1):
        cell = ws_distribution.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_distribution.column_dimensions[chr(64 + col)].width = 20
    
    for row_idx, item in enumerate(distribution_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_distribution.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    report_dir = os.path.join(settings.BASE_DIR, 'reports')
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)
    
    filename = f"用户报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(report_dir, filename)
    workbook.save(filepath)
    
    return f"用户报表已生成: {filepath}"


@app.task(name="system.tasks.generate_log_report")
def generate_log_report(days: int = 30):
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    workbook = Workbook()
    styles = create_styles(workbook)
    
    ws_operation = workbook.active
    ws_operation.title = "操作日志趋势"
    
    logs = (
        OperationLog.objects
        .filter(create_datetime__gte=start_date)
        .annotate(date=TruncDate('create_datetime'))
        .values('date', 'status')
        .annotate(count=Count('id'))
        .order_by('date')
    )
    
    date_dict = {}
    for item in logs:
        date_key = item['date']
        if date_key not in date_dict:
            date_dict[date_key] = {"success_count": 0, "fail_count": 0}
        if item['status']:
            date_dict[date_key]['success_count'] = item['count']
        else:
            date_dict[date_key]['fail_count'] = item['count']
    
    data = []
    for i in range(days):
        current_date = (start_date + timedelta(days=i)).date()
        item_data = date_dict.get(current_date, {"success_count": 0, "fail_count": 0})
        data.append({
            "日期": current_date.strftime("%Y-%m-%d"),
            "成功次数": item_data.get('success_count', 0),
            "失败次数": item_data.get('fail_count', 0),
            "总计": item_data.get('success_count', 0) + item_data.get('fail_count', 0)
        })
    
    headers = ["日期", "成功次数", "失败次数", "总计"]
    for col, header in enumerate(headers, 1):
        cell = ws_operation.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_operation.column_dimensions[chr(64 + col)].width = 18
    
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_operation.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    ws_login = workbook.create_sheet("登录日志趋势")
    
    login_logs = (
        LoginLog.objects
        .filter(create_datetime__gte=start_date)
        .annotate(date=TruncDate('create_datetime'))
        .values('date')
        .annotate(count=Count('id'))
        .order_by('date')
    )
    
    date_dict = {item['date']: item['count'] for item in login_logs}
    login_data = []
    for i in range(days):
        current_date = (start_date + timedelta(days=i)).date()
        login_data.append({
            "日期": current_date.strftime("%Y-%m-%d"),
            "登录次数": date_dict.get(current_date, 0)
        })
    
    headers = ["日期", "登录次数"]
    for col, header in enumerate(headers, 1):
        cell = ws_login.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_login.column_dimensions[chr(64 + col)].width = 20
    
    for row_idx, item in enumerate(login_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_login.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    report_dir = os.path.join(settings.BASE_DIR, 'reports')
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)
    
    filename = f"日志报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(report_dir, filename)
    workbook.save(filepath)
    
    return f"日志报表已生成: {filepath}"


@app.task(name="system.tasks.generate_org_report")
def generate_org_report():
    workbook = Workbook()
    styles = create_styles(workbook)
    
    ws_dept = workbook.active
    ws_dept.title = "部门统计"
    
    dept_data = []
    for dept in Dept.objects.all():
        user_count = Users.objects.filter(dept=dept).count()
        dept_data.append({
            "部门名称": dept.name,
            "负责人": dept.owner or "无",
            "联系电话": dept.phone or "无",
            "邮箱": dept.email or "无",
            "用户数量": user_count,
            "状态": "启用" if dept.status else "禁用"
        })
    
    headers = ["部门名称", "负责人", "联系电话", "邮箱", "用户数量", "状态"]
    for col, header in enumerate(headers, 1):
        cell = ws_dept.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_dept.column_dimensions[chr(64 + col)].width = 18
    
    for row_idx, item in enumerate(dept_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_dept.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    ws_role = workbook.create_sheet("角色统计")
    
    role_data = []
    for role in Role.objects.all():
        user_count = Users.objects.filter(role=role).count()
        role_data.append({
            "角色名称": role.name,
            "角色编码": role.code,
            "用户数量": user_count,
            "状态": "启用" if role.status else "禁用",
            "是否管理员": "是" if role.admin else "否"
        })
    
    headers = ["角色名称", "角色编码", "用户数量", "状态", "是否管理员"]
    for col, header in enumerate(headers, 1):
        cell = ws_role.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_role.column_dimensions[chr(64 + col)].width = 18
    
    for row_idx, item in enumerate(role_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_role.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    report_dir = os.path.join(settings.BASE_DIR, 'reports')
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)
    
    filename = f"组织报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(report_dir, filename)
    workbook.save(filepath)
    
    return f"组织报表已生成: {filepath}"


@app.task(name="system.tasks.generate_all_reports")
def generate_all_reports(days: int = 30):
    generate_user_report.delay(days)
    generate_log_report.delay(days)
    generate_org_report.delay()
    return "所有报表生成任务已提交"


@app.task(name="system.tasks.test_task")
def test_task():
    print('test')
