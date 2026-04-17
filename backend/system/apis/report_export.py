# -*- coding: utf-8 -*-
# @Time    : 2024/4/17
# @Author  : 数据报表导出模块
# @FileName: report_export.py
# @Software: PyCharm
import io
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse
from ninja import Router, Query
from django.db.models import Count
from django.db.models.functions import TruncDate
from system.models import Users, Role, Dept, OperationLog, LoginLog
from utils.fu_response import FuResponse

router = Router()

EXCEL_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def create_styles(workbook):
    styles = {
        'header': Font(name='微软雅黑', size=12, bold=True, color='FFFFFF'),
        'header_fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'header_border': Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        ),
        'cell': Font(name='微软雅黑', size=10),
        'cell_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'cell_border': Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        ),
    }
    return styles


def apply_header_style(cell, styles):
    cell.font = styles['header']
    cell.fill = styles['header_fill']
    cell.alignment = styles['header_alignment']
    cell.border = styles['header_border']


def apply_cell_style(cell, styles):
    cell.font = styles['cell']
    cell.alignment = styles['cell_alignment']
    cell.border = styles['cell_border']


@router.get("/export/user/trend/excel")
def export_user_trend_excel(request, days: int = 30):
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    users = (
        Users.objects
        .filter(create_datetime__gte=start_date)
        .annotate(date=TruncDate('create_datetime'))
        .values('date')
        .annotate(count=Count('id'))
        .order_by('date')
    )
    
    date_dict = {item['date']: item['count'] for item in users}
    data = []
    for i in range(days):
        current_date = (start_date + timedelta(days=i)).date()
        data.append({
            "日期": current_date.strftime("%Y-%m-%d"),
            "新增用户数": date_dict.get(current_date, 0)
        })
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "用户增长趋势"
    
    styles = create_styles(workbook)
    
    headers = ["日期", "新增用户数"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        worksheet.column_dimensions[chr(64 + col)].width = 20
    
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type=EXCEL_CONTENT_TYPE
    )
    filename = f"用户增长趋势_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@router.get("/export/user/distribution/excel")
def export_user_distribution_excel(request):
    workbook = Workbook()
    
    styles = create_styles(workbook)
    
    ws_gender = workbook.active
    ws_gender.title = "性别分布"
    
    gender_data = []
    for value, name in Users.GENDER_CHOICES:
        count = Users.objects.filter(gender=value).count()
        gender_data.append({"性别": name, "人数": count})
    
    headers = ["性别", "人数"]
    for col, header in enumerate(headers, 1):
        cell = ws_gender.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_gender.column_dimensions[chr(64 + col)].width = 15
    
    for row_idx, item in enumerate(gender_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_gender.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    ws_status = workbook.create_sheet("状态分布")
    status_data = [
        {"状态": "启用", "人数": Users.objects.filter(status=True).count()},
        {"状态": "禁用", "人数": Users.objects.filter(status=False).count()}
    ]
    
    headers = ["状态", "人数"]
    for col, header in enumerate(headers, 1):
        cell = ws_status.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_status.column_dimensions[chr(64 + col)].width = 15
    
    for row_idx, item in enumerate(status_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_status.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    ws_dept = workbook.create_sheet("部门分布")
    dept_data = []
    for dept in Dept.objects.all():
        user_count = Users.objects.filter(dept=dept).count()
        dept_data.append({"部门": dept.name, "人数": user_count})
    
    headers = ["部门", "人数"]
    for col, header in enumerate(headers, 1):
        cell = ws_dept.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        ws_dept.column_dimensions[chr(64 + col)].width = 20
    
    for row_idx, item in enumerate(dept_data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws_dept.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type=EXCEL_CONTENT_TYPE
    )
    filename = f"用户分布统计_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@router.get("/export/log/operation/excel")
def export_log_operation_excel(request, days: int = 30):
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    logs = (
        OperationLog.objects
        .filter(create_datetime__gte=start_date)
        .annotate(date=TruncDate('create_datetime'))
        .values('date', 'status')
        .annotate(count=Count('id'))
        .order_by('date')
    )
    
    date_dict = {}
    for item in logs:
        date_key = item['date']
        if date_key not in date_dict:
            date_dict[date_key] = {"success_count": 0, "fail_count": 0}
        if item['status']:
            date_dict[date_key]['success_count'] = item['count']
        else:
            date_dict[date_key]['fail_count'] = item['count']
    
    data = []
    for i in range(days):
        current_date = (start_date + timedelta(days=i)).date()
        item_data = date_dict.get(current_date, {"success_count": 0, "fail_count": 0})
        data.append({
            "日期": current_date.strftime("%Y-%m-%d"),
            "成功次数": item_data.get('success_count', 0),
            "失败次数": item_data.get('fail_count', 0),
            "总计": item_data.get('success_count', 0) + item_data.get('fail_count', 0)
        })
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "操作日志趋势"
    
    styles = create_styles(workbook)
    
    headers = ["日期", "成功次数", "失败次数", "总计"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        worksheet.column_dimensions[chr(64 + col)].width = 15
    
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type=EXCEL_CONTENT_TYPE
    )
    filename = f"操作日志趋势_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@router.get("/export/log/login/excel")
def export_log_login_excel(request, days: int = 30):
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    logs = (
        LoginLog.objects
        .filter(create_datetime__gte=start_date)
        .annotate(date=TruncDate('create_datetime'))
        .values('date')
        .annotate(count=Count('id'))
        .order_by('date')
    )
    
    date_dict = {item['date']: item['count'] for item in logs}
    data = []
    for i in range(days):
        current_date = (start_date + timedelta(days=i)).date()
        data.append({
            "日期": current_date.strftime("%Y-%m-%d"),
            "登录次数": date_dict.get(current_date, 0)
        })
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "登录日志趋势"
    
    styles = create_styles(workbook)
    
    headers = ["日期", "登录次数"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        worksheet.column_dimensions[chr(64 + col)].width = 20
    
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type=EXCEL_CONTENT_TYPE
    )
    filename = f"登录日志趋势_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@router.get("/export/dept/excel")
def export_dept_excel(request):
    data = []
    for dept in Dept.objects.all():
        user_count = Users.objects.filter(dept=dept).count()
        data.append({
            "部门名称": dept.name,
            "负责人": dept.owner or "无",
            "联系电话": dept.phone or "无",
            "邮箱": dept.email or "无",
            "用户数量": user_count,
            "状态": "启用" if dept.status else "禁用"
        })
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "部门统计"
    
    styles = create_styles(workbook)
    
    headers = ["部门名称", "负责人", "联系电话", "邮箱", "用户数量", "状态"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        worksheet.column_dimensions[chr(64 + col)].width = 18
    
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type=EXCEL_CONTENT_TYPE
    )
    filename = f"部门统计_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@router.get("/export/role/excel")
def export_role_excel(request):
    data = []
    for role in Role.objects.all():
        user_count = Users.objects.filter(role=role).count()
        data.append({
            "角色名称": role.name,
            "角色编码": role.code,
            "用户数量": user_count,
            "状态": "启用" if role.status else "禁用",
            "是否管理员": "是" if role.admin else "否"
        })
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "角色统计"
    
    styles = create_styles(workbook)
    
    headers = ["角色名称", "角色编码", "用户数量", "状态", "是否管理员"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)
        worksheet.column_dimensions[chr(64 + col)].width = 18
    
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=item[header])
            apply_cell_style(cell, styles)
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output,
        content_type=EXCEL_CONTENT_TYPE
    )
    filename = f"角色统计_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
